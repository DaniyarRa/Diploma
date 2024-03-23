import json
from openpyxl import load_workbook, Workbook
import requests
from datetime import datetime, timedelta


def fetch_and_parse_json(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Check for any HTTP errors

        data = response.json()  # Parse JSON data from the response
        return data
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
        return None


def append_to_excel(data, from_date, filename):
    if not data:
        print("No data to append.")
        return

    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Date Excitation", "Crime Title", "Hard Code", "Latitude", "Longitude"])

    for crime in data:
        sheet.append([
            from_date,
            crime["date_excitation"],
            crime["crime_title"],
            crime["hard_code"],
            crime["location"]["lat"],
            crime["location"]["lon"]
        ])

    workbook.save(filename)
    print(f"Data appended to '{filename}' successfully.")


def main(start_date, end_date, filename):
    base_url = "https://krisha.kz/ms/geodata/crime"
    params = {
        "bounds": "76.65060394531248,43.4253116907793,77.12026947265623,43.11019905275058",
        "limit": "500",
        "fields": "crime_title,hard_code,date_excitation"
    }

    start = datetime.strptime(start_date, "%d.%m.%Y")
    end = datetime.strptime(end_date, "%d.%m.%Y")
    delta = timedelta(days=1)

    while start <= end:
        from_date = start.strftime("%d.%m.%Y")
        params["from"] = from_date
        params["to"] = start.strftime("%d.%m.%Y")
        url = f"{base_url}?{requests.compat.urlencode(params)}"

        crime_data = fetch_and_parse_json(url)
        append_to_excel(crime_data, from_date, filename)

        start += delta


if __name__ == "__main__":
    start_date = "01.09.2021"
    end_date = datetime.now().strftime("%d.%m.%Y")
    filename = "crime_data.xlsx"
    main(start_date, end_date, filename)
